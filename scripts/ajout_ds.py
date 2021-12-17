# -*- coding: utf-8 -*-
"""
Created on Wed Dec 15 22:03:05 2021

@author: Xavier Pessoles
"""


## Import de bibliothèques
import openpyxl
from pathlib import Path
import sqlite3
from evaluation.class_evaluation import Evaluation
from evaluation.class_question import Question

## Paramètres 
classe = 'PSIe'
filiere = "PCSI-PSI"
annee = "2022" # Année de passage du concours
bdd = "BDD_Evaluation.db"
type_eval = "DS"
num_eval  = 1
date_eval = "16/12/2021"
dossier_notes = "Competences"
fichier_notes = "DS_N.xlsx"

evaluation = Evaluation(classe,type_eval,num_eval,date_eval)
    
def is_eval_exist(evaluation) -> str :
    """
    Renvoie -1 si l'évalation exite. 
    Renvoi l'ID si elle existe. 
    """
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    req  = evaluation.make_req_exist()
    c.execute(req)
    res = c.fetchall()
    conn.commit()
    conn.close()
    if res ==[] or res[0][0]=="" :
        return -1
    else : 
        return res[0][0]        

def add_evaluation_bdd(evaluation:Evaluation):
    id_eval = is_eval_exist(evaluation)
    if id_eval >-1 : 
        # On supprime l'évaluation
        conn = sqlite3.connect(bdd)
        c = conn.cursor()
        req = evaluation.make_req_del_eval()
        c.execute(req)
        conn.commit()
        conn.close()
        
        # On supprime les questions liées l'évaluation
        conn = sqlite3.connect(bdd)
        c = conn.cursor()
        req = evaluation.make_req_del_question(id_eval)
        
        print(req)
        c.execute(req)
        
        conn.commit()
        conn.close()
        # On supprime les questions éleves liées à l'évaluation.
        # TODO
    
  
    # On crée l'évaluation
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    req = evaluation.make_req_insertion()
    c.execute(req)
    conn.commit()
    conn.close()
    
    





def read_bareme(dossier_notes, fichier_notes) -> list:
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Questions).

    """
    id_eval = is_eval_exist(evaluation)
    
    # Lire un fichier de notes       
    xlsx_file = Path(dossier_notes, fichier_notes)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Lecture du bareme
    # Read the active sheet:
    sheet = wb_obj['Bareme']
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    
    
    ## Récupérer le nombre d'item evalués et la liste des questions
    ###############################################################
    first_row = sheet[1] 
    nb_item=0
    ligne_Q = []
    for cell in first_row:
        if cell.value!=None and 'Q' in cell.value : 
            ligne_Q.append(cell.value)
            nb_item+=1

    # Récupération des poids des questions
    poids_row = sheet[3]
    ligne_poids = []
    for cell in poids_row:
            ligne_poids.append(cell.value)

    ligne_poids=ligne_poids[2:2+nb_item]

    bareme = []


    for row in sheet.iter_rows(max_row=nb_ligne):

        ligne = []
        for cell in row:
             ligne.append(cell.value)
        
        # On vérifie que la ligne est une compétence évaluable
        if ligne[0]!=None and is_competence_evaluable(ligne[0]) :
            code_comp = ligne[0]
            valeurs = ligne[2:2+nb_item]
            if sum(valeurs)>0 :
                index_q = -1
                for i in range(len(valeurs)):
                    if valeurs[i]!=0 :
                        index_q = i
                
                # On supprime le Q
                num_ques = int(ligne_Q[index_q][1:])
                note  = valeurs[index_q]
                poids = ligne_poids[index_q]
                q = Question(id_eval,num_ques,code_comp,note,poids)
                bareme.append(q)
    return bareme


def is_competence_evaluable(comp:str) -> bool :
    """
    Renvoie True si une compétence est évaluable. 
    Elle est évaluable si elle est associée à un semestre dans la BDD

    Parameters
    ----------
    comp : str
        DESCRIPTION.

    Returns
    -------
    bool 
        DESCRIPTION.

    """
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    req = "SELECT semestre FROM competences WHERE code = '"+ \
        comp+"'"
    c.execute(req)
    res = c.fetchall()
    conn.commit()
    conn.close()
    return res[0][0]!=""

def add_bareme_bdd(bareme:list) -> None:
    """
    bareme est une liste de question
    """
    
    for question in bareme : 
        # Détermination de l'id de la compétence    
        conn = sqlite3.connect(bdd)
        c = conn.cursor()
        req = question.make_req_id_comp(filiere)
        c.execute(req)
        res = c.fetchall()
        conn.commit()
        conn.close()
        id_comp = res[0][0]
        
        conn = sqlite3.connect(bdd)
        c = conn.cursor()
        req = question.make_req_insertion(id_comp)
        c.execute(req)
        conn.commit()
        conn.close()
        
    
    


# On ajoute l'EVAL
add_evaluation_bdd(evaluation)
# On lit le bareme
bareme = read_bareme(dossier_notes, fichier_notes)
# On ajoute le bareme a la BDD
add_bareme_bdd(bareme)

# def ajout_eleves_bdd(eleves : list):
#     # On vide la table
#     conn = sqlite3.connect(bdd)
#     c = conn.cursor()
#     req = "DELETE FROM eleves WHERE classe = '"+classe+\
#               "' AND annee = '"+annee+"'"
#     c.execute(req)
#     conn.commit()
#     conn.close()  
    
#     conn = sqlite3.connect(bdd)
#     for eleve in eleves : 
#         req = eleve.make_req()
#         c = conn.cursor()
#         c.execute(req)
#         conn.commit()
#     conn.close()
    
    
# Création de l'évaluation dans la BDD

# a=is_eval_exist(type_eval, num_eval)
# print(a)
# #all_eleves  = read_bareme(dossier_notes, fichier_notes)
#ajout_eleves_bdd(all_eleves)