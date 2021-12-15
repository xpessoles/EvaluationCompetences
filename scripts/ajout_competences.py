# -*- coding: utf-8 -*-
"""
Created on Wed Dec 15 22:03:05 2021

@author: Xavier Pessoles
"""

dossier_comp = "Competences"
fichier_comp = "Competences_PCSI_PSI.xlsx"
onglet_comp  = "Competences"
filiere = 'PCSI-PSI'

## Import de bibliothèques
import pandas as pd
import openpyxl
from pathlib import Path


class Competence : 
    """ Définition d'une compétence """
    def __init__(self,filiere):
        # Code défini par le programme
        self.code = ""
        # Type : macro compétence ou compétence
        self.type = ""
        # Nom des compétences
        self.nom_long = ""
        self.nom_court = ""
        self.semestre = ""
        self.filiere = ""
    
    def creer_comp(self,ligne:list):
        # compter le nombre de None
        nb_none = 0
        for e in ligne : 
            if e==None :
                nb_none +=1
        if nb_none == 0 :
            self.type = "Compétence"
            self.code = ligne[0]
            self.nom_long = ligne[1]
            self.nom_court = ligne[2]
            self.semestre = ligne[3]
        if nb_none == 1 :
            self.type = "Macro Compétence"
            self.code = ligne[0]
            self.nom_long = ligne[1]
            self.nom_court = ligne[2]
            
            
            
def read_file_competences(dossier_comp, fichier_comp):
    # Lire un fichier de competences       
            
    xlsx_file = Path(dossier_comp, fichier_comp)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    nb_ligne = sheet.max_row
    nb_col = sheet.max_column
    
    competences = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        
        nb_none = 0
        for e in ligne : 
            if e==None :
                nb_none +=1
        if nb_none <= 1 :
            cp = Competence(filiere)
            cp.creer_comp(ligne)
            competences.append(cp)
        
    return competences

cc = read_file_competences(dossier_comp, fichier_comp)