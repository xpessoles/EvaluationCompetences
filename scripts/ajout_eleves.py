# -*- coding: utf-8 -*-
"""
Created on Wed Dec 15 22:03:05 2021

@author: Xavier Pessoles
"""
## Paramètres 
dossier_eleve = "Competences"
fichier_eleve = "Eleves_PSIe.xlsx"
classe = 'PSIe'
annee = "2022" # Année de passage du concours
bdd = "BDD_Evaluation.db"

## Import de bibliothèques
import openpyxl
from pathlib import Path
import sqlite3



class Eleve : 
    """ Définition d'un élève """
    def __init__(self,eleve,annee,classe):
        self.nom = eleve[0]
        self.prenom = eleve[1]
        self.num = eleve[2]
        self.num_ano = eleve[3]
        self.mail = eleve[4]
        self.annee = annee
        self.classe = classe
              
    
    def make_req(self):
        req = 'INSERT INTO eleves\
            (nom,prenom,num,num_ano,annee,mail,classe) \
                VALUES ("'+self.nom+'",\
                        "'+self.prenom+'",\
                        "'+str(self.num)+'",\
                        "'+str(self.num_ano)+'",\
                        "'+str(self.annee)+'",\
                        "'+self.mail+'",\
                        "'+self.classe+'" )'
                           
        return req
    
def read_file_competences(dossier_eleve, fichier_eleve) -> list:
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Competence).

    """
    # Lire un fichier de competences       
            
    xlsx_file = Path(dossier_eleve, fichier_eleve)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    eleves = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        
        nb_none = 0
        for e in ligne : 
            if e==None :
                nb_none +=1
        if nb_none ==0 :
            
            # On numérote les élèves de 00 à nn.
            if ligne[2]<10:
                ligne[2]="0"+str(ligne[2])
            print(ligne)
            eleve = Eleve(ligne,annee,classe)
            eleves.append(eleve)
        
    return eleves


def ajout_eleves_bdd(eleves : list):
    # On vide la table
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    req = "DELETE FROM eleves WHERE classe = '"+classe+\
              "' AND annee = '"+annee+"'"
    c.execute(req)
    conn.commit()
    conn.close()  
    
    conn = sqlite3.connect(bdd)
    for eleve in eleves : 
        req = eleve.make_req()
        c = conn.cursor()
        c.execute(req)
        conn.commit()
    conn.close()
    
    
    
all_eleves  = read_file_competences(dossier_eleve, fichier_eleve)
ajout_eleves_bdd(all_eleves)