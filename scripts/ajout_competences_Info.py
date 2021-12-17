# -*- coding: utf-8 -*-
"""
Created on Wed Dec 15 22:03:05 2021

@author: Xavier Pessoles
"""
## Paramètres 
dossier_comp = "Competences"
fichier_comp = "CompetencesInformatique.xlsx"
onglet_comp  = "Competences"
filiere = 'All'
discipline = "Info"
bdd = "BDD_Evaluation.db"

## Import de bibliothèques
import openpyxl
from pathlib import Path
import sqlite3

from evaluation.class_competence import Competence



    
def read_file_competences(dossier_comp, fichier_comp) -> list:
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Competence).

    """
    # Lire un fichier de competences       
            
    xlsx_file = Path(dossier_comp, fichier_comp)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    competences = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        
        nb_none = 0
        for e in ligne : 
            if e==None :
                nb_none +=1
        if nb_none <= 1 :
            cp = Competence(filiere,discipline)
            cp.creer_comp(ligne)
            competences.append(cp)
        
    return competences


def ajout_competences_bdd(competences : list):
    # On vide la table
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    req = "DELETE FROM competences WHERE filiere = '"+filiere+\
              "' AND discipline = '"+discipline+"'"
    c.execute(req)
    conn.commit()
    conn.close()  
    
    conn = sqlite3.connect(bdd)
    for competence in competences : 
        req = competence.make_req()
        c = conn.cursor()
        c.execute(req)
        conn.commit()
    conn.close()
    
    
    
all_comp  = read_file_competences(dossier_comp, fichier_comp)
ajout_competences_bdd(all_comp)