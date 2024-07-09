# -*- coding: utf-8 -*-
"""
Created on Sun Dec 19 12:54:46 2021

@author: xpess
"""
import openpyxl
from pathlib import Path
import sqlite3
import matplotlib.pyplot as plt
import codecs
import os
import shutil

from evaluation.class_competence import Competence
from evaluation.class_question import Question
from evaluation.class_evaluation import Evaluation
from evaluation.class_eleve import Eleve



##
def read_file_competences(dossier_comp, fichier_comp,filiere,discipline) -> list:
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Competence).

    """
    # Lire un fichier de competences       
            
    xlsx_file = Path(dossier_comp, fichier_comp)
    
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    competences = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        if len(row) > 5:
            print("PEUT ETRE TROP DE COLONNES DANS LE FICHIER")
            
        #print(row)
        
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        
        print(ligne)
        nb_none = 0
        for e in ligne : 
            if e==None :
                nb_none +=1
        if nb_none <= 1 :
            cp = Competence(filiere,discipline)
            cp.creer_comp(ligne)
            competences.append(cp)
        
    return competences


def ajout_competences_bdd(competences : list, filiere, discipline, bdd:str):
    # On vide la table
    req = "DELETE FROM competences WHERE filiere = '"+filiere+\
              "' AND discipline = '"+discipline+"'"
    exec_select(bdd,req)
    
    for competence in competences : 
        req = competence.make_req()
        exec_select(bdd,req)

       



def is_eval_exist(evaluation:Evaluation,bdd) -> str :
    """
    Renvoie -1 si l'évalation exite. 
    Renvoi l'ID si elle existe. 
    """
    req  = evaluation.make_req_exist()
    res = exec_select(bdd,req)

    
    if res ==[] or res[0][0]=="" :
        return -1
    else : 
        return res[0][0]        

def del_evaluation_bdd(evaluation:Evaluation,bdd):
    id_eval = is_eval_exist(evaluation,bdd)
    if id_eval >-1 : 
        # On supprime l'évaluation
        req = evaluation.make_req_del_eval()
        exec_select(bdd,req)

        
        # On supprime les questions liées l'évaluation
        req = evaluation.make_req_del_question(id_eval)
        exec_select(bdd,req)

        
        # On supprime les commentaires liés à l'évaluation
        req = evaluation.make_req_del_commentaires_eleves(id_eval)
        exec_select(bdd,req)
        
        # On supprime les questions éleves liées à l'évaluation.
        req = evaluation.make_req_del_questions_eleves(id_eval)
        exec_select(bdd,req)
        
        #  supprimer les compétences
        # On supprime les questions éleves liées à l'évaluation.
        req = evaluation.make_req_del_competences_eleves(id_eval)
        exec_select(bdd,req)
        
    

def add_evaluation_bdd(evaluation:Evaluation,bdd):
    del_evaluation_bdd(evaluation,bdd)
    # On crée l'évaluation
    req = evaluation.make_req_insertion()
    exec_select(bdd,req)

    # On garde l'id de l'évalution
    id_eval = is_eval_exist(evaluation,bdd)
    evaluation.set_id_eval(id_eval)
    
def ajout_eleves_bdd(eleves : list,bdd,classe,annee):
    # On vide la table
    req = "DELETE FROM eleves WHERE classe = '"+classe+\
              "' AND annee = '"+annee+"'"
    exec_select(bdd,req)
    
    for eleve in eleves : 
        req = eleve.make_req()
        exec_select(bdd,req)


def read_file_eleves(dossier_eleve, fichier_eleve,annee,classe) -> list:
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Competence).

    """
    # Lire un fichier de competences       
            
    xlsx_file = Path(dossier_eleve, fichier_eleve)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    eleves = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        nb_none = 0
        for e in ligne : 
            if e == None :
                nb_none +=1
        if nb_none == 0 :
            
            # On numérote les élèves de 00 à nn.
            if ligne[2]<10:
                ligne[2]="0"+str(ligne[2])
            eleve = Eleve(ligne,annee,classe)
            eleves.append(eleve)
        
    return eleves




def read_bareme(dossier_notes, fichier_notes, evaluation:Evaluation,bdd) -> list:
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Questions).

    """
    id_eval = evaluation.id_eval
    
    # Lire un fichier de notes       
    xlsx_file = Path(dossier_notes, fichier_notes)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Lecture du bareme
    # Read the active sheet:
    sheet = wb_obj['Bareme']
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    
    
    ## Récupérer le nombre d'item evalués et la liste des questions
    ###############################################################
    first_row = sheet[1] 
    nb_item=0
    ligne_Q = []
    for cell in first_row:
        if cell.value!=None and 'Q' in cell.value : 
            ligne_Q.append(cell.value)
            nb_item+=1

    # On ajoute le nombre d'itemes évalués 
    evaluation.set_nb_ques(nb_item)


    # Récupération des poids des questions
    poids_row = sheet[3]
    ligne_poids = []
    for cell in poids_row:
            ligne_poids.append(cell.value)

    ligne_poids=ligne_poids[2:2+nb_item]

    bareme = []


    for row in sheet.iter_rows(max_row=nb_ligne):
        
        ligne = []
        for cell in row:
             ligne.append(cell.value)
   
        # On vérifie que la ligne est une compétence évaluable
        if ligne[0]!=None and is_competence_evaluable(ligne[0],bdd) :
            code_comp = ligne[0]
            valeurs = ligne[2:2+nb_item]
            if sum(valeurs)>0 : # Il y a des valeurs sur la ligne donc la compétence est évaluée
                
               
                l_index_q = [] # liste des questions ou la compétence est évaluée
                for i in range(len(valeurs)):
                    if valeurs[i]!=0 :
                        l_index_q.append(i)
                
        
                for index_q in l_index_q :
                    # On supprime le Q
                    num_ques = int(ligne_Q[index_q][1:])
                    note  = valeurs[index_q]
                    poids = ligne_poids[index_q]
                    nom = ligne_Q[index_q]
                    q = Question(id_eval,num_ques,code_comp,note,poids,nom,index_q)
                    bareme.append([q,index_q])
        
    # Tri des questions dans l'ordre du sujet
    bareme = sorted(bareme, key=lambda numq: numq[1])
    bareme = [q[0] for q in bareme]
    return bareme


def is_competence_evaluable(comp:str,bdd) -> bool :
    """
    Renvoie True si une compétence est évaluable. 
    Elle est évaluable si elle est associée à un semestre dans la BDD

    Parameters
    ----------
    comp : str
        DESCRIPTION.

    Returns
    -------
    bool 
        DESCRIPTION.

    """
    req = "SELECT semestre FROM competences WHERE code = '"+ \
        comp+"'"
    res = exec_select(bdd,req)
    
    return res[0][0]!=""

def add_bareme_bdd(bareme:list,filiere, bdd) -> None:
    """
    bareme est une liste de question
    """
    index = 0
    for question in bareme : 
        # Détermination de l'id de la compétence  
        req = question.make_req_id_comp(filiere)
        res = exec_select(bdd,req)
        id_comp = res[0][0]
        
        # Ajout de la question
        req = question.make_req_insertion(id_comp)
        res = exec_select(bdd,req)
        index = index+1
        
        #Récupération de l'id de la question
        req = question.make_req_get_id() 
        res = exec_select(bdd,req)
        id_question = res[0][0]
        question.set_id_ques(id_question)
        
        
        
    
def read_notes(dossier_notes, fichier_notes, evaluation:Evaluation,bdd) -> list:
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Questions).

    """
    
    # Lire un fichier de notes       
    xlsx_file = Path(dossier_notes, fichier_notes)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Lecture du bareme
    # Read the active sheet:
    sheet = wb_obj['Notes']
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    
    notes=[]    
    
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
             ligne.append(cell.value)
        notes.append(ligne)
    
    # On récupère le nombre de lignes qui nous intéresse : 
    # de la 5ème au nombre d'élèves +5
    nb_eleves = get_nb_eleves(evaluation,bdd)
    notes = notes[4:4+nb_eleves]
    
    # On garde les lignes du nom au commentaire,
    # c'est à dire de l'index 0 à 2+nbQuesitons+1
    notes = [note[0:2+evaluation.nb_ques+1] for note in notes]    
    
    return notes

def get_nb_eleves(evaluation,bdd):
    
    classe = evaluation.classe
    annee = evaluation.annee
    req = "SELECT COUNT (id) FROM eleves WHERE "+\
        "classe='"+classe+"'"+\
        " AND annee='"+str(annee)+"'"
    
    res = exec_select(bdd,req)
    nb = res[0][0]
    return int(nb)


def add_notes_bdd(notes,evaluation:Evaluation,bareme,bdd):
    annee = evaluation.annee
    classe = evaluation.classe
    id_eval = evaluation.id_eval
    nb_quest = evaluation.nb_ques
    for note in notes :
        
        # On récupère l'id de l'élève
        num_el = note[1]
        
        req = "SELECT id from eleves WHERE"+\
            " num="+str(num_el) +\
            " AND"+" classe='"+classe +"'"+\
            " AND"+" annee="+str(annee)
        
        res = exec_select(bdd,req)
     
        id_eleve = res[0][0]

        # Pour chacune des questions on met la note
        for i in range(nb_quest):
            index_note = i+2
            id_question= bareme[i].id_ques
            n = note[index_note]
            # On ajoute la question à la BDD
            req = "INSERT INTO questions_eleves "+\
                "(id_eleve,id_eval,id_question,note_question) VALUES ("+\
                    str(id_eleve)+","+\
                    str(id_eval)+","+\
                    str(id_question)+",'"+\
                    str(n)+"')"
            
            exec_insert(bdd,req)
            
        # Pour chaque éval, on met le commentaire
        comment = note[-1]
        req = "INSERT INTO commentaires_eleves "+\
            "(id_eleve,id_eval,commentaire) VALUES ("+\
                str(id_eleve)+","+\
                str(id_eval)+',"'+\
                comment+'")'
        
        exec_insert(bdd,req)
    
       
def exec_insert(bdd,req):
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    c.execute(req)
    conn.commit()
    conn.close()
    
def exec_select(bdd,req):
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    c.execute(req)
    res = c.fetchall()
    conn.commit()
    conn.close()
    return res

def get_eleves(classe:str,annee:int,bdd) -> list:
    """
    Liste des élèves sous forme d'Eleve
    pour une classe donnée et pour une année donnée. 
    Le tuple est ainsi constiués 
    id_bdd,nom,prenom,num,num_ano,annee,classe,mail

    Parameters
    ----------
    classe : str
        DESCRIPTION.
    annee : int
        DESCRIPTION.
    bdd : TYPE
        DESCRIPTION.

    Returns
    -------
    Liste d'élèves.

    """
    req = "SELECT id,nom,prenom,num,num_ano,annee,classe,mail FROM eleves WHERE"+\
        " annee="+str(annee)+\
        " AND classe ='"+classe+"'"+" ORDER BY num"
    res = exec_select(bdd,req)
    liste_eleves = []
    for ligne in res :
        eleve = Eleve.from_sql(ligne)
        liste_eleves.append(eleve)
    return liste_eleves 

def get_questions_eleve(evaluation,eleve:Eleve,bdd):
    id_eval = is_eval_exist(evaluation,bdd)

    #Synthèse d'un élève
    req = "SELECT id_eleve,id_eval,id_question,note_question FROM questions_eleves"+\
        " WHERE id_eval="+str(id_eval)+" AND id_eleve="+str(eleve.id)+" ORDER BY id_question"
    res = exec_select(bdd,req)
    questions_eleve = []
    for ligne in res :
        q_el = {"id_eleve":ligne[0],
                "id_eval":ligne[1],
                "id_question":ligne[2],
                "note_question":ligne[3]}
        questions_eleve.append(q_el)
    return questions_eleve
    
def get_questions_eval(id_eval,bdd):
    """
    Récupération du barème à partir de la BDD.

    Parameters
    ----------
    id_eval : TYPE
        DESCRIPTION.
    bdd : TYPE
        DESCRIPTION.

    Returns
    -------
    liste_questions : list(Question)

    """
    req = "SELECT nom,num_ques,index_question,poids_comp,"+\
          " note_ques,code FROM questions "+\
          "JOIN competences ON questions.id_comp=competences.id "+\
          " WHERE id_eval="+str(id_eval)+" ORDER BY questions.id"
    res = exec_select(bdd,req)
    liste_questions = []
    for quest in res : 
        question = Question.from_tuple(quest,id_eval)
        liste_questions.append(question)
    return liste_questions

def calc_note_eval(bareme,notes_eleve):
    id_eleve = notes_eleve[0]["id_eleve"]
    note_brute = 0
    total_brut = 0
    note_traitee = 0
    total_traite = 0
    #print(">>>>",notes_eleve)
    for i in range(len(bareme)):
        note_quest = notes_eleve[i]['note_question']
        poids_question = bareme[i].poids
        note_quest_bareme = bareme[i].note
        
        if note_quest!="NT":
            try :
                x = (float(note_quest)*note_quest_bareme)
            except ValueError : 
            
                print("Value Error : ", note_quest,note_quest_bareme,id_eleve)
            
            note_traitee += float(note_quest)*note_quest_bareme
            total_traite += note_quest_bareme*poids_question
        else : 
            note_quest = 0
           
        note_brute += float(note_quest)*note_quest_bareme
        total_brut += note_quest_bareme*poids_question
    
    note_brute = note_brute*20/total_brut
    if total_traite > 0 : # Si l'élève est présent
        note_traitee = note_traitee*20/total_traite
    else : 
        note_traitee = 0
    return {"note_brute":note_brute,"total_brut":total_brut,\
            "note_traitee":note_traitee,"total_traite":total_traite,\
            "id_eleve":id_eleve}

def insert_note_eval_bdd(id_eleve,id_eval,note_eval_eleve,bdd):
    req = "DELETE FROM evaluations_eleves WHERE  "+\
        "id_eleve = '"+ str(id_eleve)+\
        "' AND id_evaluation = "+ str(id_eval)
    exec_select(bdd,req)

    req = "INSERT INTO evaluations_eleves "+\
        "(id_eleve,id_evaluation,note_brute_sur20,note_traitee_sur20) VALUES ("+\
            str(id_eleve)+","+\
            str(id_eval)+","+\
            str(note_eval_eleve['note_brute'])+","+\
            str(note_eval_eleve['note_traitee'])+")"
    
    exec_select(bdd,req)
    
    
def insert_comp_bdd(eleve,id_eval,notes_eleve,bareme,filiere,bdd):
    
    id_eleve = eleve.id
    req = "DELETE FROM competences_eleves WHERE  "+\
         "id_eleve = '"+ str(id_eleve)+\
         "' AND id_evaluation = "+ str(id_eval)
    exec_select(bdd,req)
    
    
    
    for i in range(len(bareme)):
        Ques = bareme[i]
        req = Ques.make_req_id_comp(filiere)
        id_comp = exec_select(bdd,req)[0][0]
        
        score = 'NT'
        poids_ques = bareme[i].poids
        note = notes_eleve[i]["note_question"]
        if note=='NT':
            score = 'NT'
        else : 
            score = float(note)*100/poids_ques
        
        
        req = "INSERT INTO competences_eleves "+\
            "(id_eleve,id_evaluation,id_comp,score) VALUES ("+\
                str(id_eleve)+","+\
                str(id_eval)+","+\
                str(id_comp)+",'"+\
                str(score)+"')"
        
        exec_select(bdd,req)
        
def classement_eval(bilan_evals):
    # Ajouter le classement de l'évaluation
    
    # note_eval_eleve : liste de dictionnaire 
    #{'note_brute': 0.00, 'total_brut': 105, 'note_traitee': 9.0, 
    # 'total_traite': 100, 'id_eleve': 156}
    
    # Conversion de la liste de dico en liste de liste pour 
    # pouvoir faire un tri
    liste_evals =  sorted(bilan_evals, key=lambda evals: evals['note_brute'], reverse=True)
    for i in range(0,len(liste_evals)):
        dico = liste_evals[i]
        dico["Rang_brut"]=i+1
        liste_evals[i]=dico
    
    liste_evals =  sorted(bilan_evals, key=lambda evals: evals['note_traitee'], reverse=True)
    for i in range(0,len(liste_evals)):
        dico = liste_evals[i]
        dico["Rang_traite"]=i+1
        liste_evals[i]=dico
        
    return liste_evals

def plot_notes_brute(id_eleve,bilan_evals,fichier):
    # Graphique des notes avec positionnement de l'élève
    bilan_evals =  sorted(bilan_evals, key=lambda evals: evals['note_brute'], reverse=True)
    
    # On cherche le rang et la note brutes de id_eleve
    for eleve in bilan_evals : 
        if eleve["id_eleve"] == id_eleve :
            note = eleve["note_brute"]
            rang = eleve["Rang_brut"]
    les_notes = [note["note_brute"] for note in bilan_evals]
    les_rang = [note["Rang_brut"] for note in bilan_evals]
    # Toutes les notes
    plt.plot(les_rang,les_notes,"b.") 
    # Les notes de l'éleve
    plt.plot(rang,note,"rs")
    plt.xlabel("Classement")
    plt.ylabel("Notes")
    plt.grid()
    plt.savefig(fichier)
    plt.close()
        

def calc_moyenne_classe(liste_evals):
    moyenne_brute=[]
    for l in liste_evals:
        moyenne_brute.append(l["note_brute"])
    return sum(moyenne_brute)/len(moyenne_brute)

def ecriture_notes_eleves_tex(eleve,notes_eleve,id_eval,bareme,liste_evals,file_el,bdd,coef_ds,ord_origine):
    """
    Création des spécificités du .tex spécifiques à l'évaluation. 
    """
    id_eleve = eleve.id
    nom = eleve.nom
    prenom = eleve.prenom
    for l in liste_evals : 
        if l["id_eleve"]==id_eleve :
            note_brute = l["note_brute"]
            rang_brut = l["Rang_brut"]
            note_traitee = l["note_traitee"]
            Rang_traite = l["Rang_traite"]
    moyenne_classe = calc_moyenne_classe(liste_evals)
    
    
    # Récupération des commentaires perso
    req = "SELECT commentaire FROM commentaires_eleves WHERE "+\
        " id_eleve="+str(id_eleve)+" AND"+\
        " id_eval="+str(id_eval)
    
    # Recup de la note par competences du DS
    eval_comp_ds = eval_competences_ds(bareme,bdd)
    eval_comp_ds = sorted(eval_comp_ds.items(), key=lambda t: t[0])
    
    
    commentaire = exec_select(bdd, req)[0][0]
    fid = codecs.open(file_el, "w", "utf-8")

    nb_questions = len(notes_eleve)
    
    # ===== EN TETE ELEVE ====
    
    fid.write("\\begin{minipage}[c]{.45\\linewidth} \n")
    
    fid.write("\\Large \\textbf{\\textsf{"+nom.upper()+" "+prenom+"}} \n \n")  
    
    fid.write(" \\normalsize Note brute "+str(round(note_brute,2))+"/20 \n \n")
    
    fid.write(" \\normalsize Note harmonisée "+str(round(coef_ds*note_brute+ord_origine,2))+"/20 \n \n")
    fid.write("Rang "+str(rang_brut)+"\n \n")
    #fid.write("Note brute "+str(round(bilan_el[1],2))+"/20 \n \n")
    
    fid.write("Moyenne classe brute "+str(round(moyenne_classe,2))+"/20 \n \n")

    
    fid.write("Moyenne question traitées "+str(round(note_traitee,2))+"/20 \n \n")
    fid.write("Rang question traitées "+str(Rang_traite)+" \n \n")
    
    fid.write("Commentaires : \n")
    fid.write(commentaire+" \n")
    fid.write("\\end{minipage}\\hfill \n")
    fid.write("\\begin{minipage}[c]{.45\\linewidth}  \n")
    fid.write("\\begin{center}\n")
    fid.write("\\includegraphics[width=.8\\linewidth]{../histo.pdf} \n")
    fid.write("\\end{center}\n")
    
    fid.write("\\end{minipage}\n")
    
    

    
    # ===== NOTES PAR QUESTIONS =====
    # On ajoute les notes par questions
    fid.write("\\footnotesize \n")
    fid.write("\\begin{center} \n")
    fid.write("\\begin{tabular}{|c|c|m{1cm}|c||c|c|m{1cm}|c||c|c|m{1cm}|c||c|c|m{1cm}|c|} \n")
    fid.write("\\hline "+\
        "\\textbf{Qu} & \\textbf{Coef} & \\textbf{Comp} & \\textbf{/5} & "+\
        "\\textbf{Qu} & \\textbf{Coef} & \\textbf{Comp} & \\textbf{/5} & "+\
        "\\textbf{Qu} & \\textbf{Coef} & \\textbf{Comp} & \\textbf{/5} & "+\
        "\\textbf{Qu} & \\textbf{Coef} & \\textbf{Comp} & \\textbf{/5} \\\ \n")
    fid.write("\\hline \n")
    fid.write("\\hline \n")
    
    
    # On traite les n-1 premières lignes
    nb_lignes = (nb_questions-1)//4
    for i in range (0,nb_lignes):
        col1 = 4*i
        col2 = 4*i+1
        col3 = 4*i+2
        col4 = 4*i+3
        
        numques_1 = bareme[col1].nom
        numques_2 = bareme[col2].nom
        numques_3 = bareme[col3].nom
        numques_4 = bareme[col4].nom
        
        #coef/poids
        coef_1 = str(bareme[col1].note)
        coef_2 = str(bareme[col2].note)
        coef_3 = str(bareme[col3].note)
        coef_4 = str(bareme[col4].note)
        
        # compétences
        comp_1 = bareme[col1].code_comp
        comp_2 = bareme[col2].code_comp
        comp_3 = bareme[col3].code_comp
        comp_4 = bareme[col4].code_comp
        
        # notes
        note_1 = notes_eleve[col1]["note_question"]
        note_2 = notes_eleve[col2]["note_question"]
        note_3 = notes_eleve[col3]["note_question"]
        note_4 = notes_eleve[col4]["note_question"]
        
        
        ligne = numques_1+" & "+coef_1+" & "+comp_1+" & "+note_1+" & "+\
            numques_2+" & "+coef_2+" & "+comp_2+" & "+note_2+" & "+\
            numques_3+" & "+coef_3+" & "+comp_3+" & "+note_3+" & "+\
            numques_4+" & "+coef_4+" & "+comp_4+" & "+note_4+" \\\ \\hline \n "
              
       
        fid.write(ligne)
        fid.write('\n')
        
    
    # On écrit la dernière ligne
    #############################
    col1 = 4*nb_lignes
    col2 = 4*nb_lignes+1
    col3 = 4*nb_lignes+2
    col4 = 4*nb_lignes+3
    
    numques_1, numques_2, numques_3, numques_4 = "","","",""
    coef_1, coef_2, coef_3, coef_4 = "","","",""
    comp_1,comp_2,comp_3,comp_4 = "","","",""
    note_1,note_2,note_3,note_4 = "","","",""
    
    
    if col1+1 <=nb_questions :
        numques_1 = bareme[col1].nom
        coef_1 = str(bareme[col1].note)
        comp_1 = bareme[col1].code_comp    
        note_1 = notes_eleve[col1]["note_question"]
    if col1+2 <=nb_questions :
        numques_2 = bareme[col2].nom
        coef_2 = str(bareme[col2].note)
        comp_2 = bareme[col2].code_comp    
        note_2 = notes_eleve[col2]["note_question"]
    if col1+3 <=nb_questions :
        numques_3 = bareme[col3].nom
        coef_3 = str(bareme[col3].note)
        comp_3 = bareme[col3].code_comp    
        note_3 = notes_eleve[col3]["note_question"]
    if col1+4 <=nb_questions :
        numques_4 = bareme[col4].nom
        coef_4 = str(bareme[col4].note)
        comp_4 = bareme[col4].code_comp    
        note_4 = notes_eleve[col4]["note_question"]
        
    ligne = numques_1+" & "+coef_1+" & "+comp_1+" & "+note_1+" & "+\
        numques_2+" & "+coef_2+" & "+comp_2+" & "+note_2+" & "+\
        numques_3+" & "+coef_3+" & "+comp_3+" & "+note_3+" & "+\
        numques_4+" & "+coef_4+" & "+comp_4+" & "+note_4+" \\\ \\hline \n "
    fid.write(ligne)
    fid.write('\n')
    
    ############################## FIN DERNIERE LIGNE ##########
    
    fid.write("\\end{tabular} \n")
    fid.write("\\end{center} \n")
    fid.write("\\normalsize \n \n")
    # ===== FIN NOTES PAR QUESTIONS =====
    
    
    #### Bilan de compétence sur le DS. 
    
    fid.write("\\noindent \\textbf{Bilan par compétences}\n \n")
    
    fid.write("\\begin{itemize} \n")
    for comp in eval_comp_ds :
        r = comp[1]["note"]/comp[1]["note_total"]
        r = str(int(100*r))
        fid.write("\\item \\textbf{"+comp[0]+"} : "+comp[1][comp[0]]+ "\\hfill Réussite : " + r +"\\% \n")
    fid.write("\\end{itemize} \n")    
    fid.close()

def generation_bilan_eval_indiv(classe,annee,filiere,evaluation,bdd,coef_ds,ord_origine,ext):
    # ext : extension pour le fichier PDF (par exemple s'il y a un sujet 
    # CCMP:"_CCMP", un sujet CCS :"_CCS")
    # Récupération des élèves

    eleves = get_eleves(classe,annee,bdd)
    id_eval = is_eval_exist(evaluation,bdd)

    bilan_evals = []

    #print(len(eleves))
    for eleve in eleves : 


        # Récup du bareme Liste de Questions
        bareme = get_questions_eval(id_eval,bdd)

        # Récupération des notes d'un élève
        # Dictionnaire de notes d'un éleve
        notes_eleve = get_questions_eleve(evaluation,eleve,bdd)
        # Calcul de l'élève
        #print(bareme,notes_eleve)
        note_eval_eleve = calc_note_eval(bareme,notes_eleve)
        # On écrit ca dans la base de données
        insert_note_eval_bdd(eleve.id,id_eval,note_eval_eleve,bdd)
        
        # On ajoute les compétences évaluées dans la base.
        insert_comp_bdd(eleve,id_eval,notes_eleve,bareme,filiere,bdd)
        
        bilan_evals.append(note_eval_eleve)

    # Ajouter le classement de l'évaluation
    liste_evals = classement_eval(bilan_evals)
    
    
    for eleve in eleves :
        
        # Ecriture fichier tex
        print(eleve.nom)
        notes_eleve = get_questions_eleve(evaluation,eleve,bdd)
        
        ecriture_notes_eleves_tex(eleve,notes_eleve,id_eval,bareme,liste_evals,"compil/f1.tex",bdd,coef_ds,ord_origine)
        plot_notes_brute(eleve.id,bilan_evals,"compil/histo.pdf")
        os.chdir("compil")
        
        print(' >> ICI << ')
        os.system("pdflatex FicheDS.tex")
        os.system("pdflatex FicheDS.tex")
        fichier_eleve = eleve.get_num()+"_"+\
                        eleve.nom+"_"+\
                        eleve.prenom+"_"+\
                        evaluation.type_eval+"_"+\
                        str(evaluation.num_eval)+ext+".pdf"
        shutil.move("FicheDS.pdf",fichier_eleve)
        os.chdir("..")
        

def get_comp_id(code_comp,discipline, filiere, bdd):
    # Récupération de l'id d'une compétence
    req = "SELECT id from competences WHERE "+\
        "code = '"+code_comp +"' AND "+\
        "discipline = '"+discipline +"' AND "+\
        "filiere = '"+filiere+"'"
    res = exec_select(bdd,req)
    return res[0][0]

def get_liste_score_eleve(id_eleve,id_comp, bdd):
    # Récupération du type d'éval, de la date, du score
    req = "SELECT evaluations.type,evaluations.date,score FROM "+\
    "competences_eleves JOIN evaluations ON "+\
    "evaluations.id = competences_eleves.id_evaluation "+\
    " WHERE id_eleve = "+str(id_eleve)+\
    " AND id_comp = "+str(id_comp)+\
    " order by evaluations.id"
    #print("ici IL Y A UN PROBLEME : on recupere pas tout ce qu'on veut")
    res = exec_select(bdd,req)
    return res

def plot_score_comp(scores,filename):
    notes, dates, type_eval = [], [], []
    notes_DS, dates_DS = [], []
    notes_colles, dates_colles = [], []
    notes_DDS, dates_DDS = [], []
    # scores: liste de score
    # score : (type,date,score)
    for score in scores :
        type_eval.append(score[0])
        dates.append(score[1])
        if score[2]=="NT" : 
            nnn = 0
        else :
            nnn = float(score[2])
        notes.append(nnn)
        
        if score[0] == "DS": 
            dates_DS.append(score[1])
            notes_DS.append(nnn)
        if score[0] == "DDS": 
            dates_DDS.append(score[1])
            notes_DDS.append(nnn)
        if score[0] == "colle": 
            dates_colles.append(score[1])
            notes_colles.append(nnn)
    
    
    labels = dates
    val = notes
    width = 0.01      # the width of the bars: can also be len(x) sequence

    fig, ax = plt.subplots(figsize=(6.5,1.5))


    ax.bar(labels, val, width)

    ax.bar(dates_DS, notes_DS, width, label='DS',color="cornflowerblue")
    ax.bar(dates_colles, notes_colles, width, label='Colles',color="cornflowerblue")
    ax.bar(dates_DDS, notes_DDS, width, label='DDS',color="cornflowerblue")


    ax.plot(labels, val,color="cornflowerblue",linewidth=1)
    ax.scatter(labels, val,color="royalblue",s=20)
    ax.set_ylim(0, 100)
    
    
    
    #ax.set_yticks([0, 25, 50, 75, 100])
    ax.set_xticklabels(dates,fontsize=6)
    ax.set_yticklabels([0, 25, 50, 75, 100],fontsize=7)
    #ax.set_xticklabels(fontsize=6)
    ax.set_ylabel('Scores',fontsize=6)
    ax.legend(prop={'size': 6})
    plt.savefig(filename)

def generation_bilan_competences(eleve,classe,filiere,discipline,bdd) :
    req = "SELECT code,nom_long FROM competences WHERE "+\
        "discipline = '"+discipline+\
        "' AND filiere = '"+filiere+\
        "' order by id"
    id_eleve = eleve.id
    res = exec_select(bdd,req)
    fid = codecs.open("compil/comp.tex", "w", "utf-8")

    for line in res : 
        code = line[0]
        nom = line[1]
        if len(code)==1 : 
            titre = "\section{"+code+" -- "+nom+"}  \n"
            fid.write(titre)
            #print(titre)
        elif len(code)==2 : 
            titre = "\subsection{"+code+" -- "+nom+"}  \n"
            fid.write(titre)
            #print(titre)
        else :
            titre = "\subsubsection*{"+code+" -- "+nom+"}  \n"
            fid.write(titre)
            id_comp = get_comp_id(code,discipline, filiere, bdd)
            
            scores = get_liste_score_eleve(id_eleve,id_comp, bdd)
            
            if len(scores)>0:
                plot_score_comp(scores,"compil/"+code+".pdf")
                fid.write("\\begin{center} \n")
                fid.write("\\includegraphics{"+code+".pdf} \n")
                fid.write("\\end{center} \n")
                
    fid.close()

def eval_competences_ds(bareme,bdd):
    """
    Pour un DS donné : 
        - faire la liste des compétences évaluées
        - pour chaque compétence : 
                - trouver le nom long
                - donner la note de l'élève
                - donner la la note totale 
        On renvout un dico de dico :
         { code_comp:{'nom':nom long,'note':int, 'note_tot':int}}
    """
    eval_comp_ds = {}
    for q in bareme :
        if q.code_comp not in eval_comp_ds : 
            req = "SELECT nom_long from competences WHERE"+\
                " code='"+str(q.code_comp)+"'"
            res = exec_select(bdd,req)
            d={q.code_comp:res[0][0],"note":0,"note_total":0} 
            eval_comp_ds[q.code_comp] = d
            
    for q in bareme :
        eval_comp_ds[q.code_comp]['note']+= q.note
        eval_comp_ds[q.code_comp]['note_total']+= q.poids
    
    return eval_comp_ds
    
            
         