# -*- coding: utf-8 -*-
"""
Created on Thu Jan 13 13:10:17 2022

@author: xpess
"""

# Lecture d'un fichier de note d'écrit
import xlrd     # fichiers xls
import sqlite3  
import openpyxl # fichiers xlsx
from pathlib import Path



file = "2021_PSI_Etoile_CentraleSupelec_Oral.xls"
file_ccinp = "2012_PSI_Etoile_CCINP_Ecrit.xls"
file_banque = "Ecoles.xlsx"
dossier_banque = ""
bdd = "StatConcours.db"
ecole="CCINP"
classe = "PSI_Etoile"
annee = "2021"
filiere = "PSI"

def lecture_banques(dossier:str, file:str) : #-> list[dict]:
    # Lire un fichier de notes       
    xlsx_file = Path(dossier, file)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Lecture du bareme
    # Read the active sheet:
    sheet = wb_obj['Banque']
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    banques = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        banques.append(ligne[0])
    return banques

def lecture_ecoles(dossier:str, file:str) : #-> list[dict]:
    # Lire un fichier de notes       
    xlsx_file = Path(dossier, file)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Lecture du bareme
    # Read the active sheet:
    sheet = wb_obj['Ecoles']
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    ecoles = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        ecole={"banque":ligne[0],"ecole":ligne[1],'filiere':ligne[2],\
               "inscription":ligne[3],"places":ligne[4]}
        ecoles.append(ecole)
    return ecoles[1:]
    
    
    
def lecture_notes(file:str) : #-> list[dict]:
    """
    A partir d'un fichier xls provenant de scei, renvoie une liste de dictionnaires
    Chaque dictionnaire etant constitué des notes des eleves d'une classe

    Parameters
    ----------
    file : str
        Fichier xls.

    Returns
    -------
    list[dict]
        DESCRIPTION.

    """
    
    myBook = xlrd.open_workbook(file)
    sheet = myBook.sheet_by_index(0)
    
    # Caractéristiques du concours
    carac = sheet.cell_value(2, 0) 
    carac = carac.split("   ")
    CONCOURS = carac[0]
    BARRE = carac[1]
    
    if "barre" in BARRE : 
        BARRE=BARRE.split(' ')[1]
        
    # Caractéristiques d'une ligne (en vue d'en faire un dictionnaire)
    titre_colonnes = sheet.row(3)
    titre_colonnes = [v.value for v in titre_colonnes]
    
    dico_notes = []
    # Création d'une liste de dictionnaires
    for i in range(4,sheet.nrows):
        eleve = {}
        ligne = sheet.row(i)
        ligne = [v.value for v in ligne]
        eleve["ecole"]=CONCOURS
        for j in range(len(ligne)):
            eleve[titre_colonnes[j]]=ligne[j]
        dico_notes.append(eleve)
    return dico_notes


def exec_req(bdd,req):
    conn = sqlite3.connect(bdd)
    c = conn.cursor()
    c.execute(req)
    res = c.fetchall()
    conn.commit()
    conn.close()
    return res

def is_eleve_existe(num_scei:int,annee:int,bdd:str) -> bool:
    """
    Verfie si un eleve existe dans la BDD
    """
    req = "SELECT id FROM eleves WHERE "+\
        "num_scei='"+str(num_scei)+"'" +\
        " AND annee='"+str(annee)+"'"
    res = exec_req(bdd,req)
    
    if res ==[] or res[0][0]=="" :
        return False
    else : 
        return True

def is_eleve_inscrit_ecole(id_eleve:int,id_ecole:int,bdd:str) -> bool:
    """
    Verfie si un eleve est inscrit dans une école
    """
    req = "SELECT * FROM eleve_ecole WHERE "+\
        "id_eleve='"+str(id_eleve)+"'" +\
        " AND id_ecole='"+str(id_ecole)+"'"
    res = exec_req(bdd,req)
    
    if res ==[] or res[0][0]=="" :
        return False
    else : 
        return True
    
def is_banque_existe(banque:str,bdd:str) -> bool:
    """
    Verfie qu'un banque existe dans la bdd
    """
    req = "SELECT id FROM banques WHERE "+\
        "nom='"+banque+"'"
    res = exec_req(bdd,req)
    
    if res ==[] or res[0][0]=="" :
        return False
    else : 
        return True

def is_ecole_existe(ecole:dict,bdd:str) -> bool:
    """
    Verfie qu'une école existe dans la bdd
    """
    req = "SELECT id FROM ecoles WHERE "+\
        'nom="'+ecole['ecole']+'"'
    res = exec_req(bdd,req)

    if res ==[] or res[0][0]=="" :
        return None
    else : 
        return res[0]


def get_id_eleve(num_scei:int,annee:int,bdd:str) -> bool:
    """
    Verfie si un eleve existe dans la BDD
    """
    req = "SELECT id FROM eleves WHERE "+\
        "num_scei='"+str(num_scei)+"'" +\
        " AND annee='"+str(annee)+"'"
    res = exec_req(bdd,req)
    
    return res[0][0]
    
def inscription_eleve(id_eleve,id_ecole,bdd:str) -> bool:
    """
    Inscrit élève
    """
    req = "INSERT INTO 'eleve_ecole' (id_eleve,id_ecole) VALUES " +\
        "("+str(id_eleve)+","+str(id_ecole)+")"
    exec_req(bdd,req)
    


# def creation_eleve_bdd(eleve:dict,redoublant,classe,annee,bdd:str) -> None:
#     req = "INSERT INTO eleves "+\
#         "(num_scei,nom_prenom,classe,redoublant,annee) VALUES ("+\
#             eleve["N°"]+","+\
#             eleve["Nom"]+","+\
#             classe+",'"+\
#             redoublant+",'"+\
#             str(annee)+"')"
#     exec_req(bdd,req)
    
# def ecrire_notes_bdd(bdd,dico_notes,classe,annee):
    
#     for eleve in dico_notes : 
#         num_scei = eleve["N°"]
#         # On regarde si l'eleve existe, sinon on le crée
#         if not(is_eleve_existe(num_scei,bdd)):
#             # Par défaut on met les élèves en 3/2
#             creation_eleve_bdd(eleve, "Non", classe, annee, bdd)
#         id_eleve = get_id_eleve(num_scei,annee)
#         # On inscrit l'élève aux écoles, si c'est des notes d'écrit
#         # Pour CCMP, CCINP, CCMT, quand on inscrit aux banques, on inscrit à toutes
#         # les écoles de la banque 
#         if eleve["ecole"]=='Concours Commun INP PSI':
#             banque = "CCINP"

def ajout_banque(banques,bdd):
    """
    banques : liste des banques
    Parameters
    ----------
    banques : TYPE
        DESCRIPTION.
    bdd : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    for banque in banques : 
        if not(is_banque_existe(banque, bdd)):
            req = "INSERT INTO banques "+\
            "(nom) VALUES ('"+\
                str(banque)+"')"
          
            exec_req(bdd,req)

def get_banque_id(ecole,bdd):
    """
    Parameters
    ----------
    ecole : TYPE
        DESCRIPTION.
    bdd : TYPE
        DESCRIPTION.

    Returns
    -------
    res : TYPE
        DESCRIPTION.

    """
    if type(ecole)==dict :
        banque = ecole["banque"]
    else : 
        banque = ecole
        
    req = "SELECT id FROM banques WHERE "+\
        "nom='"+banque+"'"
        
    res = exec_req(bdd,req)
    return res[0][0]


def get_liste_ecole_id(banque,bdd):
    """
    banque : int
    Récupération de la liste des écoles d'une banque dont l'inscription se fait par banque
    (Dans CCINP pour l'inscription est directe pour certains ecoles (CPE Lyon par exemple) 
     ou par ecole pour d'autres (ESM par exemple))
    """
    

        
    req = "SELECT id FROM ecoles WHERE "+\
        "id_banque="+str(banque)+\
        " AND inscription='banque'"
        
    res = exec_req(bdd,req)
    res = [e[0] for e in res]
    return res



def add_ecole_bdd(ecole,id_banque,bdd):
    
    if ecole["places"]==None : ecole["places"]=-1
    
    req = "INSERT INTO ecoles "+\
    '(nom,id_banque,inscription,filiere,places) VALUES ("'+\
        ecole["ecole"]+'","'+\
        str(id_banque)+'","'+\
        ecole["inscription"]+'","'+\
        ecole["filiere"]+'",'+\
        str(ecole["places"])+')'        
    res = exec_req(bdd,req)
    return res

def add_eleve_bdd(note,classe,redoublant,annee):
    num_scei = int(note["N°"])
    nom_prenom = note["Nom"]
    if not(is_eleve_existe(num_scei, annee, bdd)):
            
        req = "INSERT INTO eleves "+\
        '(num_scei,nom_prenom,classe,redoublant,annee) VALUES ("'+\
            str(num_scei)+'","'+\
            nom_prenom+'","'+\
            classe+'","'+\
            redoublant+'",'+\
            annee+')'        
    res = exec_req(bdd,req)
    return res

def ajout_ecoles(ecoles,bdd):
    """

    Parameters
    ----------
    ecoles : Liste de dictionnaires de la forme 
    {'banque': 'CCINP',
     'ecole': 'ESM Saint-Cyr',
     'filiere': 'PSI',
     'inscription': 'ecole',
     'places': None}
    
    bdd : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    for ecole in ecoles :
        # On vérifie si l'école existe dans la BDD
        id_ecole = is_ecole_existe(ecole, bdd)
        if id_ecole == None : 
            # SI l'ecole n'existe pas on cherche l'id de la banque associée
            banque_id = get_banque_id(ecole,bdd)[0][0]
            add_ecole_bdd(ecole, banque_id, bdd)

def ecriture_notes_bdd(notes:list,ecole:str,classe:str,filiere:str,annee:int,bdd):
    """
    ecole :nom de la banque ou de l'école suivant le fichier de notes
    """
    
    for note in notes : 
        # on regarde si l'eleve existe
        num_scei = note["N°"]
        eleve_id = is_eleve_existe(num_scei, annee, bdd)
        
        # Si l'eleve n'existe pas, on l'inscrit
        if eleve_id == False : 
            # Par défaut on le met 3/2
            redoublant = "non"
            add_eleve_bdd(note,classe,redoublant,annee)
        
        eleve_id = get_id_eleve(num_scei, annee, bdd)
        
        # on regarde si la note existe
        print(eleve_id)
        
        # Si l'école est une banque, on l'inscrit à toutes les écoles de la banque
        if ecole=="CCMP" or ecole == "CCMT" or ecole == "CCINP":
            # On récupèr l'id de la banque
            id_banque = get_banque_id(ecole, bdd)
            print(id_banque)
            
            # On récupère toutes les écoles de la banque en question
            # ou l'inscription se fait par école
            liste_ecoles = get_liste_ecole_id(id_banque, bdd)
            for id_ecole in liste_ecoles :
                if not(is_eleve_inscrit_ecole(eleve_id, id_ecole, bdd)) :
                    inscription_eleve(eleve_id, id_ecole, bdd)
            
                # On ajoute les notes si ca n'a pas été fait
                if not(is_*****(eleve_id, id_ecole, bdd)) :
                    add_notes_eleve_ccinp_ecrit(eleve_id, id_ecole, bdd)
                    
           
            
        
        
    return None
    
    
banques = lecture_banques(dossier_banque,file_banque)
ajout_banque(banques, bdd)
ecoles = lecture_ecoles(dossier_banque,file_banque)
ajout_ecoles(ecoles, bdd)
notes = lecture_notes(file_ccinp)
ecriture_notes_bdd(notes, ecole, classe, filiere, annee, bdd)